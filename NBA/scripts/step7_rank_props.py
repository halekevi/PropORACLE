#!/usr/bin/env python3
"""
step7_rank_props.py  (VECTORIZED 2026-03-01)

PERF: All 12 .apply() calls and 2 row-by-row list comprehensions replaced with
      vectorized pandas/NumPy operations. Estimated 3-4x faster on 8,000+ row slates.
      Excel write engine switched from openpyxl → xlsxwriter (~5x faster write).

PATCH (2026-02-26):
- Fix edge_adj_dr to be direction-aware: UNDERs now get a positive edge
  contribution when projection < line.
- Support projection building for volume props (2PTA/2PTM, 3PTA/3PTM, FTA/FTM, FGA/FGM).
- Adds prop_norm aliases for volume props.

PATCH (2026-02-23):
- grading-informed reweight of scoring components.
"""

from __future__ import annotations

import argparse
import numpy as np
import pandas as pd

# UTF-8 safe Excel export
try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

# -------------------- helpers --------------------

def _to_num(s):
    return pd.to_numeric(s, errors="coerce")

def _norm_pick_type_series(s: pd.Series) -> pd.Series:
    t = s.astype(str).str.strip().str.lower()
    return np.where(t.str.contains("gob"), "Goblin",
           np.where(t.str.contains("dem"), "Demon", "Standard"))

# -------------------- weights --------------------

# Prop weights — calibrated from 9-day graded outcomes (2026-03-06 → 2026-03-14)
# Higher weight = model gives this prop more scoring influence.
# Lowered props that were systematically over-predicted; raised fantasy (best predictor).
_PROP_WEIGHTS = {
    "fantasy":             1.08,  # 65.3% actual HR — best predictor, was 1.00
    "pts":                 1.03,  # 58.1% — slight lower, was 1.03
    "pr":                  1.03,  # 56.6% — was 1.01
    "reb":                 1.02,  # 55.3% — was 1.06
    "ra":                  1.02,  # 54.9% — was 1.02
    "pra":                 1.03,  # 54.8% — was 0.99
    "pa":                  1.02,  # 54.2% — was 1.01
    "ast":                 1.01,  # 53.0% — was 1.05
    "fg2a":                1.03,  # 52.8% — was 0.92 (minor raise)
    "fga":                 1.03,  # 52.2% — was 0.99
    "pf":                  0.97,  # 52.0% — was 0.85 (raise — undervalued)
    "personalfouls":       0.97,
    "tov":                 0.96,  # 51.6% — was 0.94
    "fgm":                 0.95,  # 51.2% — was 0.99
    "fg3a":                0.97,  # 50.9% — was 0.88 (minor raise)
    "3ptattempted":        0.97,
    "twopointersattempted":0.98,
    "ftm":                 0.94,  # 50.2% — was 1.01
    "freethrowsmade":      0.94,
    "stocks":              0.92,  # 49.2% — was 1.04
    "stl":                 0.92,  # 49.1% — was 1.08 (big drop)
    "fg2m":                0.95,  # 49.1% — was 1.01
    "twopointersmade":     0.95,
    "fta":                 0.94,  # 48.6% — was 0.98
    "freethrowsattempted": 0.94,
    "fg3m":                0.90,  # 47.8% — was 1.03 (big drop)
    "3ptmade":             0.90,
    "blk":                 0.80,  # 38.7% — worst OVER prop, was 1.02
}

# Hit rate priors — calibrated from 9-day graded data (2026-03-06 → 2026-03-14)
# Used in prop_hr_z scoring signal. Old values were based on season-long prior;
# these reflect actual pipeline output hit rates by prop type OVER direction.
_PROP_HR_PRIOR_OVER = {
    "fantasy":             0.700,  # actual 68.8%, was 0.674
    "pts":                 0.580,  # actual 59.4%, was 0.566
    "pr":                  0.565,  # actual 57.6%, was 0.568
    "reb":                 0.580,  # actual 56.1%, was 0.617 (lowered)
    "ra":                  0.555,  # actual 54.8%, was 0.600 (lowered)
    "ast":                 0.555,  # actual 53.2%, was 0.593 (lowered)
    "fga":                 0.510,  # actual 50.0%, was 0.558 (lowered)
    "pra":                 0.545,  # actual 54.9%, was 0.545
    "pa":                  0.550,  # actual 54.8%, was 0.557
    "fgm":                 0.510,  # actual 50.8%, was 0.519
    "fg2m":                0.510,  # actual 49.1%, was 0.528
    "twopointersmade":     0.510,
    "fg2a":                0.520,  # actual 52.8%, unchanged
    "twopointersattempted":0.520,
    "tov":                 0.500,  # actual 49.8%, was 0.484 (slight raise)
    "pf":                  0.510,  # actual 52.0%, was 0.424 (significant raise)
    "personalfouls":       0.510,
    "fg3a":                0.490,  # actual 51.3%, was 0.444 (raise)
    "3ptattempted":        0.490,
    "stocks":              0.510,  # actual 48.4%, was 0.547 (lowered)
    "stl":                 0.530,  # actual 48.1%, was 0.697 (major drop — UNDER is the signal)
    "fg3m":                0.520,  # actual 47.0%, was 0.623 (major drop)
    "3ptmade":             0.520,
    "ftm":                 0.510,  # actual 48.0%, was 0.583 (lowered)
    "freethrowsmade":      0.510,
    "fta":                 0.460,  # actual 42.0%, was 0.545 (lowered)
    "freethrowsattempted": 0.460,
    "blk":                 0.420,  # actual 38.3%, was 0.545 (major drop)
}

# UNDER overrides — calibrated from 9-day graded data.
# These are props where the UNDER signal is meaningfully different from (1 - OVER prior).
# Key insight: Steals/3PM/FTA/Blks+Stls UNDER are the best Standard signals in the dataset.
_PROP_HR_PRIOR_UNDER_OVERRIDE = {
    "stl":                 0.667,  # actual 66.7% UNDER — was 0.303, massive raise
    "fg3m":                0.580,  # actual 60.0% UNDER — was 0.377
    "3ptmade":             0.580,
    "stocks":              0.580,  # actual 60.0% UNDER — was 0.453
    "fta":                 0.580,  # actual 59.3% UNDER — was 0.455
    "freethrowsattempted": 0.580,
    "ra":                  0.540,  # actual 55.7% UNDER — was 0.400
    "ftm":                 0.545,  # actual 55.2% UNDER — was 0.417
    "freethrowsmade":      0.545,
    "ast":                 0.515,  # actual 50.4% UNDER — was 0.407
    "tov":                 0.590,  # actual 60.4% UNDER — was 0.516
    "fga":                 0.545,  # actual 54.0% UNDER — was 0.645 (lowered)
    "fg2a":                0.545,  # actual 48.4% UNDER — was 0.645 (lowered)
    "twopointersattempted":0.480,
    "reb":                 0.495,  # actual 49.5% UNDER — was 0.591 (lowered)
    "pa":                  0.515,  # actual 51.2% UNDER — was 0.590 (lowered)
    "pts":                 0.492,  # actual 49.2% UNDER — was 0.540 (lowered)
    "pr":                  0.506,  # actual 50.6% UNDER — was 0.540 (lowered)
    "pra":                 0.535,  # actual 53.5% UNDER — was 0.540
    "fantasy":             0.288,  # actual 28.8% UNDER — hard near-block, was 0.371
    "pf":                  0.518,  # actual 51.8% UNDER — was derived
    "personalfouls":       0.518,
}

_RELIABILITY_MAP = {
    "Standard": 1.00,  # baseline
    "Goblin":   1.08,  # was 1.06 — consistently outperforms, slight raise
    "Demon":    0.50,  # was 0.75 — 31.8% actual hit rate, needs to be near-invisible
}

# -------------------- projection fallback --------------------

_PLAYER_PREFIX_BY_PROP = {
    "fga": "fga", "fgm": "fgm", "fg2a": "fg2a", "fg2m": "fg2m",
    "fg3a": "fg3a", "fg3m": "fg3m", "fta": "fta", "ftm": "ftm",
}

_COMBO_CORRECTIONS = {"pr": 1.05, "pa": 1.06, "ra": 1.08, "pra": 1.07, "fantasy": 1.15}

def _edge_transform_series(edge: pd.Series, cap: float = 3.0, power: float = 0.85) -> pd.Series:
    """Vectorized power-transform with sign preservation."""
    sign = np.sign(edge)
    clipped = np.clip(edge.abs(), 0, cap)
    return sign * (clipped ** power)

def _tier_from_score_series(score: pd.Series) -> pd.Series:
    return np.where(score >= 2.50, "A",
           np.where(score >= 1.75, "B",
           np.where(score >= 1.10, "C", "D")))

def _write_xlsx_openpyxl(output_path: str, out: pd.DataFrame, elig_mask: pd.Series) -> None:
    """Write XLSX with explicit UTF-8 encoding using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create both sheets with UTF-8 safe values
    for sheet_name, df_sheet in [("ALL", out), ("ELIGIBLE", out.loc[elig_mask])]:
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(df_sheet, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                # Ensure value is properly UTF-8 encoded (especially for player names)
                if isinstance(value, str):
                    # Force string through UTF-8 encode/decode to ensure proper handling
                    value = value.encode('utf-8').decode('utf-8')
                elif pd.isna(value):
                    value = None
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Set encoding in workbook properties
    wb.properties.encoding = 'UTF-8'
    wb.save(output_path)
    print(f"✅ Saved → {output_path} (openpyxl, UTF-8 encoded)")

# -------------------- main --------------------

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  default="step6_with_team_role_context.csv")
    ap.add_argument("--output", default="step7_ranked_props.xlsx")
    args = ap.parse_args()

    print(f"→ Loading: {args.input}")
    df = pd.read_csv(args.input, dtype=str, encoding="utf-8-sig", 
                     engine='python').fillna("")
    
    # Explicitly ensure all string columns are str type (not object with mixed types)
    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].astype(str)
    out = df.copy()

    for col, default in [("line", ""), ("pick_type", "Standard"), ("prop_norm", "")]:
        if col not in out.columns:
            out[col] = default

    if "prop_norm" not in out.columns or out["prop_norm"].eq("").all():
        if "prop_type" in out.columns:
            out["prop_norm"] = out["prop_type"].astype(str).str.lower()

    # Normalize prop names
    _PROP_NORM_MAP = {
        "3-pt made": "fg3m", "3-pt attempted": "fg3a",
        "3pt made": "fg3m", "3pt attempted": "fg3a",
        "three pointers made": "fg3m", "three pointers attempted": "fg3a",
        "3-ptm": "fg3m", "3-pta": "fg3a", "3ptm": "fg3m", "3pta": "fg3a",
        "two pointers made": "fg2m", "two pointers attempted": "fg2a",
        "2 pointers made": "fg2m", "2 pointers attempted": "fg2a",
        "2pt made": "fg2m", "2pt attempted": "fg2a",
        "2-pt made": "fg2m", "2-pt attempted": "fg2a",
        "2-ptm": "fg2m", "2-pta": "fg2a", "2ptm": "fg2m", "2pta": "fg2a",
        "free throws made": "ftm", "free throws attempted": "fta",
        "freethrowsmade": "ftm", "freethrowsattempted": "fta",
        "ft made": "ftm", "ft attempted": "fta", "ftm": "ftm", "fta": "fta",
        "fg attempted": "fga", "fg made": "fgm",
        "field goals attempted": "fga", "field goals made": "fgm",
        "fga": "fga", "fgm": "fgm",
        "fg3a": "fg3a", "fg3m": "fg3m", "fg2a": "fg2a", "fg2m": "fg2m",
    }
    out["prop_norm"] = (out["prop_norm"].astype(str).str.lower().str.strip()
                        .map(lambda x: _PROP_NORM_MAP.get(x, x)))

    prop_norm_s = out["prop_norm"].astype(str).str.lower().str.strip()
    line_num    = _to_num(out["line"])
    pick_type_s = pd.Series(_norm_pick_type_series(out["pick_type"]), index=out.index)

    # ── VECTORIZED PROJECTION ─────────────────────────────────────────────────
    v5  = _to_num(out.get("stat_last5_avg",  ""))
    v10 = _to_num(out.get("stat_last10_avg", ""))
    vs  = _to_num(out.get("stat_season_avg", ""))

    # Weighted blend (50/30/20) with partial weight normalization
    w5 = np.where(v5.notna(),  0.50, 0.0)
    w10= np.where(v10.notna(), 0.30, 0.0)
    ws = np.where(vs.notna(),  0.20, 0.0)
    total_w = w5 + w10 + ws
    proj_raw = (
        v5.fillna(0)  * w5 +
        v10.fillna(0) * w10 +
        vs.fillna(0)  * ws
    )
    proj_raw = np.where(total_w > 0.1, proj_raw / total_w, np.nan)

    # Fallback for volume props: look for {prefix}_player_last5_avg etc.
    missing_proj = np.isnan(proj_raw)
    if missing_proj.any():
        for prop_key, prefix in _PLAYER_PREFIX_BY_PROP.items():
            mask = missing_proj & (prop_norm_s == prop_key)
            if not mask.any():
                continue
            for col_cand in [f"{prefix}_player_last5_avg", f"{prefix}_last5_avg"]:
                if col_cand in out.columns:
                    fb = _to_num(out[col_cand])
                    proj_raw = np.where(mask & fb.notna(), fb, proj_raw)
                    break

    # Combo/fantasy correction
    corr = prop_norm_s.map(lambda x: _COMBO_CORRECTIONS.get(x, 1.0)).values
    proj = pd.Series(proj_raw * corr, index=out.index)
    out["projection"] = proj

    out["edge"]     = proj - line_num
    out["abs_edge"] = out["edge"].abs()
    # Normalized edge keeps cross-prop comparisons on the same scale
    # (e.g., Fantasy Score vs Points).
    line_safe = line_num.replace(0, np.nan)
    out["edge_norm"] = out["edge"] / line_safe

    # ── FORCED OVER / BET DIRECTION ───────────────────────────────────────────
    forced = pick_type_s.isin(["Goblin", "Demon"]).astype(int)
    out["forced_over_only"] = forced

    bet_dir = np.where(forced.eq(1), "OVER",
              np.where(_to_num(out["edge"]) >= 0, "OVER", "UNDER"))
    out["bet_direction"] = bet_dir

    # ── ELIGIBILITY ───────────────────────────────────────────────────────────
    miss       = line_num.isna() | proj.isna()
    # Goblin/Demon with negative edge: drop to audit sheet, exclude from scoring
    neg_forced = forced.eq(1) & (_to_num(out["edge"]) < 0)
    drop_mask  = neg_forced  # rows that go to DROPPED tab only

    eligible    = (~miss & ~drop_mask).astype(int)
    void_reason = pd.Series("", index=out.index)
    void_reason = void_reason.where(~miss,      "NO_PROJECTION_OR_LINE")
    void_reason = void_reason.where(~drop_mask, "DROPPED_NEG_EDGE_GOBDEM")

    # ── HARD BLOCKS: prop+direction combinations with <45% hit rate on Standard ──
    # Derived from 9-day calibration (2026-03-06 → 2026-03-14, 19,461 props).
    # These are blocked regardless of edge — the model has no predictive power here.
    _BLOCKED_STD_OVER = {
        "stl", "blk",               # Steals OVER 41.9%, Blocks OVER 38.3%
        "fta", "freethrowsattempted",# FT Attempted OVER 43.9%
        "stocks",                    # Blks+Stls OVER 47.5% (marginal, block for safety)
    }
    _BLOCKED_STD_UNDER = {
        "fantasy",                   # Fantasy Score UNDER 28.8% — worst prop in dataset
        "pts",                       # Points UNDER 49.2% — coin flip, not worth Standard slot
    }
    _BLOCKED_ANY_UNDER = {
        "reb",                       # Rebounds UNDER 49.5% overall (13.3% some days)
    }

    is_standard = pick_type_s == "Standard"
    is_over     = pd.Series(bet_dir, index=out.index) == "OVER"
    is_under    = ~is_over

    block_std_over  = is_standard & is_over  & prop_norm_s.isin(_BLOCKED_STD_OVER)
    block_std_under = is_standard & is_under & prop_norm_s.isin(_BLOCKED_STD_UNDER)
    block_any_under = is_under & prop_norm_s.isin(_BLOCKED_ANY_UNDER)
    hard_block      = block_std_over | block_std_under | block_any_under

    eligible    = np.where(hard_block, 0, eligible)
    void_reason = pd.Series(
        np.where(block_std_over,  "BLOCKED_STD_OVER_LOW_HR",
        np.where(block_std_under, "BLOCKED_STD_UNDER_LOW_HR",
        np.where(block_any_under, "BLOCKED_UNDER_LOW_HR",
        void_reason))),
        index=out.index
    )
    eligible    = pd.Series(eligible, index=out.index)

    # ── PER-PROP MINIMUM EDGE THRESHOLDS for Standard OVER ───────────────────
    # Props need higher edge to be actionable on Standard (no line discount).
    # Thresholds derived from edge bins where Standard OVER first crosses 55% HR.
    _STD_OVER_MIN_EDGE = {
        "pts":   2.0,   # Points OVER needs +2 minimum on Standard
        "pa":    3.0,   # Pts+Asts OVER very noisy — needs +3
        "pra":   2.5,   # Pts+Rebs+Asts OVER — needs +2.5
        "pr":    2.0,   # Pts+Rebs OVER — needs +2
        "ra":    1.5,   # Rebs+Asts OVER — needs +1.5
        "ast":   1.5,   # Assists OVER — needs +1.5
        "fg3m":  1.5,   # 3-PT Made OVER — needs +1.5
        "3ptmade":1.5,
    }
    for prop_k, min_e in _STD_OVER_MIN_EDGE.items():
        edge_too_low = (
            is_standard & is_over
            & (prop_norm_s == prop_k)
            & (_to_num(out["edge"]) < min_e)
        )
        eligible    = np.where(edge_too_low, 0, eligible)
        void_reason = pd.Series(
            np.where(edge_too_low, f"STD_OVER_EDGE_BELOW_MIN_{prop_k.upper()}", void_reason),
            index=out.index
        )
    eligible    = pd.Series(eligible, index=out.index)

    out["eligible"]    = eligible
    out["void_reason"] = void_reason

    elig_mask = eligible.eq(1)

    # ── VECTORIZED EDGE TRANSFORM ─────────────────────────────────────────────
    out["edge_dr"] = _edge_transform_series(_to_num(out["edge_norm"]))

    # ── VECTORIZED LINE HIT RATE ──────────────────────────────────────────────
    # Direction-aware: pick the right column priority
    bet_is_under = pd.Series(bet_dir, index=out.index) == "UNDER"

    def _pick_first_valid(*col_names) -> pd.Series:
        result = pd.Series(np.nan, index=out.index)
        for col in col_names:
            if col in out.columns:
                v = _to_num(out[col])
                result = result.where(result.notna(), v)
        return result

    hr5_over  = _pick_first_valid("line_hit_rate_over_ou_5",  "line_hit_rate_over_5",  "last5_hit_rate")
    hr10_over = _pick_first_valid("line_hit_rate_over_ou_10", "line_hit_rate_over_10")
    hr5_under = _pick_first_valid("line_hit_rate_under_ou_5", "line_hit_rate_under_5")
    hr10_under= _pick_first_valid("line_hit_rate_under_ou_10","line_hit_rate_under_10")

    # Derived under from counts if direct column missing
    l5o = _to_num(out.get("last5_over",  ""))
    l5u = _to_num(out.get("last5_under", ""))
    denom_ou = (l5o + l5u).replace(0, np.nan)
    derived_under = l5u / denom_ou
    hr5_under = hr5_under.where(hr5_under.notna(), derived_under)

    # No push fallback (1 - over) when push==0
    l5p = _to_num(out.get("last5_push", ""))
    hr5_under = hr5_under.where(hr5_under.notna(),
        np.where(l5p.fillna(0) == 0, 1.0 - hr5_over, np.nan))

    hr5  = np.where(bet_is_under, hr5_under, hr5_over)
    hr10 = np.where(bet_is_under, hr10_under, hr10_over)
    hr5  = pd.Series(hr5,  index=out.index)
    hr10 = pd.Series(hr10, index=out.index)

    # Blend 5 and 10 game windows
    line_hit_rate = (
        np.where(hr5.notna() & hr10.notna(), hr5 * 0.50 + hr10 * 0.50,
        np.where(hr5.notna(),  hr5,
        np.where(hr10.notna(), hr10, np.nan)))
    )
    out["line_hit_rate"] = pd.Series(line_hit_rate, index=out.index)

    # ── VECTORIZED MINUTES CERTAINTY ──────────────────────────────────────────
    _MIN_TIER_MAP = {"HIGH": 1.00, "MEDIUM": 0.90, "LOW": 0.75}
    out["minutes_certainty"] = (
        out.get("minutes_tier", pd.Series("", index=out.index))
        .astype(str).str.upper()
        .map(lambda x: _MIN_TIER_MAP.get(x, 0.80))
    )

    # ── VECTORIZED PROP WEIGHT / RELIABILITY ─────────────────────────────────
    out["prop_weight"]      = prop_norm_s.map(lambda x: _PROP_WEIGHTS.get(x, 0.93))
    out["reliability_mult"] = pick_type_s.map(lambda x: _RELIABILITY_MAP.get(x, 0.97))

    # ── VECTORIZED DEF ADJUSTMENT ─────────────────────────────────────────────
    def_rank = _to_num(out.get("OVERALL_DEF_RANK", ""))
    def_adj  = ((def_rank - 15.0) / 15.0 * 0.06).fillna(0.0)
    out["def_adj"] = def_adj

    # ── GAME CONTEXT ADJUSTMENT (Step 6b: Vegas lines) ────────────────────────
    # ctx_adj: -0.08 low total on combo prop, -0.05 blowout risk, -0.15 both
    ctx_adj  = _to_num(out["ctx_adj"]).fillna(0.0)  if "ctx_adj"  in out.columns else pd.Series(0.0, index=out.index)
    out["ctx_adj"] = ctx_adj

    # ── SCHEDULE / REST ADJUSTMENT (Step 6c: B2B, rest days) ─────────────────
    # rest_adj: -0.10 B2B, 0.00 baseline (1-day rest), +0.02 two days, +0.04 three+
    rest_adj = _to_num(out["rest_adj"]).fillna(0.0) if "rest_adj" in out.columns else pd.Series(0.0, index=out.index)
    out["rest_adj"] = rest_adj

    # ── PACE SIGNAL ──────────────────────────────────────────────────────────
    # Derived from game_total (Step 6b). High total = fast pace = more possessions.
    # Neutral at 230pts, ±0.02 per 10pt deviation, capped ±0.04.
    # Direction-aware: fast pace helps OVER props, hurts UNDER props.
    if "game_total" in out.columns:
        pace_raw   = (_to_num(out["game_total"]).fillna(230.0) - 230.0) / 10.0 * 0.02
        pace_adj   = pace_raw.clip(-0.04, 0.04)
        pace_adj_dr = pd.Series(
            np.where(bet_is_under, -pace_adj, pace_adj), index=out.index
        )
    else:
        pace_adj_dr = pd.Series(0.0, index=out.index)
    out["pace_adj"] = pace_adj_dr

    # ── PROP-SPECIFIC OPP ALLOWANCE ───────────────────────────────────────────
    # intel_opp_vs_league_pct (Step 6e) measures how much more/less this
    # opponent gives up vs league avg for the specific stat being scored.
    # +8% on an AST prop = opponent gives up 8% more assists = stronger OVER signal.
    # Converted to a projection multiplier: ±0.02 per 5% deviation, capped ±0.06.
    # This is separate from the general intel_def_z weight (0.40) in the score —
    # it directly adjusts the projection so edge and hit rates benefit too.
    opp_pct_raw    = _to_num(out.get("intel_opp_vs_league_pct", pd.Series(np.nan, index=out.index))).fillna(0.0) / 100.0
    opp_prop_adj   = (opp_pct_raw / 0.05 * 0.02).clip(-0.06, 0.06)
    opp_prop_adj_dr = pd.Series(
        np.where(bet_is_under, -opp_prop_adj, opp_prop_adj), index=out.index
    )
    out["opp_prop_adj"] = opp_prop_adj_dr

    proj_base = _to_num(out["projection"])
    out["projection_adj"] = proj_base * (
        1.0 + def_adj + ctx_adj + rest_adj + pace_adj_dr + opp_prop_adj_dr
    )
    out["edge_adj"]       = out["projection_adj"] - line_num
    out["edge_adj_norm"]  = out["edge_adj"] / line_safe

    # ── VECTORIZED EDGE_ADJ_DR (direction-aware) ──────────────────────────────
    edge_adj_signed = np.where(bet_is_under, -_to_num(out["edge_adj_norm"]), _to_num(out["edge_adj_norm"]))
    out["edge_adj_dr"] = _edge_transform_series(pd.Series(edge_adj_signed, index=out.index))

    # ── VECTORIZED DEF RANK SIGNAL ────────────────────────────────────────────
    signal_raw = ((def_rank - 1.0) / 29.0 * 2.0 - 1.0)
    def_signal = np.where(bet_is_under, -signal_raw, signal_raw)
    out["def_rank_signal"] = pd.Series(def_signal, index=out.index)

    # ── VECTORIZED PROP HIT RATE PRIOR ───────────────────────────────────────
    base_prior = prop_norm_s.map(lambda x: _PROP_HR_PRIOR_OVER.get(x, 0.545))
    under_prior = prop_norm_s.map(
        lambda x: _PROP_HR_PRIOR_UNDER_OVERRIDE.get(x, 1.0 - _PROP_HR_PRIOR_OVER.get(x, 0.545))
    )
    out["prop_hr_prior"] = np.where(bet_is_under, under_prior, base_prior)

    # ── VECTORIZED AVG VS LINE ────────────────────────────────────────────────
    for col in ("stat_last5_avg", "stat_last10_avg", "stat_season_avg"):
        out[col + "_num"] = _to_num(out[col]) if col in out.columns else pd.Series(np.nan, index=out.index)

    def _avg_vs_line_vec(avg_col: str, w: float) -> pd.Series:
        v = _to_num(out[avg_col + "_num"]) if (avg_col + "_num") in out.columns else pd.Series(np.nan, index=out.index)
        raw = np.clip((v - line_safe) / line_safe, -1.0, 1.0)
        raw = np.where(bet_is_under, -raw, raw)
        return pd.Series(np.where(v.notna() & line_safe.notna(), raw * w, np.nan), index=out.index)

    avl5  = _avg_vs_line_vec("stat_last5_avg",  0.50)
    avl10 = _avg_vs_line_vec("stat_last10_avg", 0.30)
    avls  = _avg_vs_line_vec("stat_season_avg", 0.20)

    wt5  = np.where(_to_num(out.get("stat_last5_avg_num",  "")).notna() & line_safe.notna(), 0.50, 0.0)
    wt10 = np.where(_to_num(out.get("stat_last10_avg_num", "")).notna() & line_safe.notna(), 0.30, 0.0)
    wts  = np.where(_to_num(out.get("stat_season_avg_num", "")).notna() & line_safe.notna(), 0.20, 0.0)
    total_avl_w = pd.Series(wt5 + wt10 + wts, index=out.index)

    avg_vs_line = (avl5.fillna(0) + avl10.fillna(0) + avls.fillna(0))
    avg_vs_line = avg_vs_line.where(total_avl_w > 0.1, 0.0)
    out["avg_vs_line"] = avg_vs_line

    # ── Z-SCORE (direction-aware) ─────────────────────────────────────────────
    def zcol(s: pd.Series, direction_aware: bool = False) -> pd.Series:
        x = pd.to_numeric(s, errors="coerce")
        result = pd.Series(0.0, index=x.index)
        if direction_aware and "bet_direction" in out.columns:
            for direction in ("OVER", "UNDER"):
                dir_mask = elig_mask & (out["bet_direction"].astype(str).str.upper() == direction)
                if dir_mask.sum() < 2:
                    continue
                mu = x[dir_mask].mean()
                sd = x[dir_mask].std()
                if sd and not np.isnan(sd) and sd > 1e-9:
                    z_vals = (x[dir_mask] - mu) / sd
                    result.loc[dir_mask.index[dir_mask]] = z_vals.values
            return result
        mu = x[elig_mask].mean()
        sd = x[elig_mask].std()
        if sd and not np.isnan(sd) and sd > 1e-9:
            return (x - mu) / sd
        return result

    out["edge_z"]        = zcol(out["edge_norm"],      direction_aware=True)
    out["line_hit_z"]    = zcol(out["line_hit_rate"],   direction_aware=True)
    out["min_z"]         = zcol(out["minutes_certainty"])
    out["def_rank_z"]    = zcol(out["def_rank_signal"],  direction_aware=True)
    out["avg_vs_line_z"] = zcol(out["avg_vs_line"],      direction_aware=True)
    out["prop_hr_z"]     = zcol(out["prop_hr_prior"],    direction_aware=True)

    # ── Intel signals (from step6e) ───────────────────────────────────────────
    # intel_season_hit_rate: % of season games OVER this line (0-100 scale → normalise)
    intel_shr_raw  = _to_num(out.get("intel_season_hit_rate", pd.Series(np.nan, index=out.index))).fillna(50.0) / 100.0
    # intel_opp_vs_league_pct: how generous/tight this opponent is (+= give up more)
    intel_def_raw  = _to_num(out.get("intel_opp_vs_league_pct", pd.Series(np.nan, index=out.index))).fillna(0.0) / 100.0
    # intel_cv_pct: consistency — lower = better. Invert so high = consistent
    intel_cv_raw   = _to_num(out.get("intel_cv_pct", pd.Series(np.nan, index=out.index))).fillna(50.0)
    intel_cons_raw = (100.0 - intel_cv_raw.clip(0, 100)) / 100.0  # 0-1, higher=consistent

    out["intel_shr_z"]  = zcol(pd.Series(intel_shr_raw,  index=out.index), direction_aware=True)
    out["intel_def_z"]  = zcol(pd.Series(intel_def_raw,  index=out.index), direction_aware=True)
    out["intel_cons_z"] = zcol(pd.Series(intel_cons_raw, index=out.index))

    # ── FINAL SCORE ───────────────────────────────────────────────────────────
    # Scoring weight rationale (from 9-day calibration, 19,461 decided props):
    #
    # SIGNAL            CORR    OLD WT  NEW WT  NOTES
    # edge_adj_dr       0.132   0.85    0.90    signed edge — best single predictor
    #                                           now also bakes in pace + opp_prop_adj
    # intel_shr_z       —       0.70    0.85    season hit rate at this line — strong
    # line_hit_z        0.148   0.85    0.80    recent hit rate — reliable
    # avg_vs_line_z     —       0.75    0.70    avg vs line — useful but noisy
    # def_rank_z        —       0.80    0.60    overall defense rank
    # prop_hr_z         —       0.50    0.55    calibrated priors
    # intel_def_z       —       0.40    0.35    opp generosity — lowered: now also in
    #                                           projection via opp_prop_adj (double-count risk)
    # intel_cons_z      —       0.25    0.25    consistency — unchanged
    # pace_z            —       NEW     0.20    game pace signal (from game_total)
    # min_z             —       0.25    0.20    minutes certainty — minor
    #
    # Context penalties (flat, not z-scored):
    # b2b_penalty       -0.20   on player's own team B2B flag
    # blowout_penalty   -0.10   blowout risk games
    # low_total_pen     -0.10   low total games

    b2b_penalty    = np.where(out.get("b2b_flag",    pd.Series(False, index=out.index)).astype(str).str.lower() == "true", -0.20, 0.0)
    blowout_penalty= np.where(out.get("blowout_risk", pd.Series(False, index=out.index)).astype(str).str.lower() == "true", -0.10, 0.0)
    low_total_pen  = np.where(out.get("low_total_flag", pd.Series(False, index=out.index)).astype(str).str.lower() == "true", -0.10, 0.0)

    # Pace z-score (direction-aware — already in out["pace_adj"] but score via z for scale)
    out["pace_z"] = zcol(out["pace_adj"], direction_aware=True)

    score = (
        _to_num(out["edge_adj_dr"]).fillna(0.0)      * 0.90   # best predictor; now includes pace+opp
        + _to_num(out["intel_shr_z"]).fillna(0.0)   * 0.85   # season hit rate at this line
        + _to_num(out["line_hit_z"]).fillna(0.0)     * 0.80   # recent hit rate
        + _to_num(out["avg_vs_line_z"]).fillna(0.0)  * 0.70   # avg vs line
        + _to_num(out["def_rank_z"]).fillna(0.0)     * 0.60   # overall defense rank
        + _to_num(out["prop_hr_z"]).fillna(0.0)      * 0.55   # calibrated priors
        + _to_num(out["intel_def_z"]).fillna(0.0)   * 0.35   # lowered: opp allowance now in proj
        + _to_num(out["intel_cons_z"]).fillna(0.0)  * 0.25   # consistency
        + _to_num(out["pace_z"]).fillna(0.0)         * 0.20   # NEW: game pace signal
        + _to_num(out["min_z"]).fillna(0.0)          * 0.20   # minutes certainty
        + pd.Series(b2b_penalty,     index=out.index)
        + pd.Series(blowout_penalty, index=out.index)
        + pd.Series(low_total_pen,   index=out.index)
    )
    score = (
        score
        * _to_num(out["prop_weight"]).fillna(1.0)
        * _to_num(out["reliability_mult"]).fillna(1.0)
    )
    score = score.where(elig_mask, np.nan)

    out["rank_score"] = score
    out["tier"] = pd.Series(
        _tier_from_score_series(_to_num(out["rank_score"])), index=out.index
    )
    out.loc[~elig_mask, "tier"] = "D"

    # Split here — after all scoring/tier columns are populated
    dropped_df = out.loc[drop_mask].copy()
    out_active = out.loc[~drop_mask].copy()

    # ── WRITE XLSX (with explicit UTF-8 handling) ──────────────────────────────
    # Sheets:
    #   ALL        — all active rows (neg-edge Gob/Dem excluded)
    #   STANDARD   — Standard pick type only
    #   GOB_DEM    — Goblin + Demon (positive-edge only)
    #   ELIGIBLE   — active rows that passed scoring
    #   DROPPED    — neg-edge Goblin/Demon, for hit/miss audit only

    std_mask_active  = out_active["pick_type"].astype(str).str.strip().str.lower().str.contains("standard")
    gobdem_mask      = ~std_mask_active
    elig_mask_active = out_active["eligible"].eq(1)

    def _safe_excel_write(writer, df, sheet_name):
        if df.empty:
            pd.DataFrame(columns=df.columns).to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    if HAS_XLSXWRITER:
        try:
            with pd.ExcelWriter(args.output, engine="xlsxwriter",
                               engine_kwargs={'options': {'strings_to_urls': False}}) as w:
                _safe_excel_write(w, out_active,                        "ALL")
                _safe_excel_write(w, out_active.loc[std_mask_active],   "STANDARD")
                _safe_excel_write(w, out_active.loc[gobdem_mask],       "GOB_DEM")
                _safe_excel_write(w, out_active.loc[elig_mask_active],  "ELIGIBLE")
                _safe_excel_write(w, dropped_df,                        "DROPPED")
            print(f"✅ Saved → {args.output} (xlsxwriter, UTF-8 encoded)")
        except Exception as e:
            print(f"⚠️  xlsxwriter failed: {e}, falling back to openpyxl")
            _write_xlsx_openpyxl(args.output, out_active, elig_mask_active)
    else:
        _write_xlsx_openpyxl(args.output, out_active, elig_mask_active)

    print(f"✅ Saved → {args.output}")
    print(f"ALL rows (active) : {len(out_active)}")
    print(f"STANDARD rows     : {int(std_mask_active.sum())}")
    print(f"GOB_DEM rows      : {int(gobdem_mask.sum())}")
    print(f"DROPPED rows      : {len(dropped_df)}  (neg-edge Gob/Dem, audit only)")
    print()
    print("Tier counts (ALL active):")
    print(out_active["tier"].value_counts().to_string())
    print()
    print("Ineligible reason breakdown (active):")
    vr = out_active.loc[~elig_mask_active, "void_reason"].value_counts()
    print(vr.to_string() if len(vr) else "(none)")
    print()
    print("Score percentiles (eligible):")
    rs = _to_num(out_active.loc[elig_mask_active, "rank_score"])
    print(rs.quantile([0.50, 0.70, 0.80, 0.85, 0.90, 0.95]).round(3).to_string())


if __name__ == "__main__":
    main()
